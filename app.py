import re
import unicodedata
import warnings
import io

import pandas as pd
import plotly.express as px
import streamlit as st

warnings.filterwarnings("ignore")

MUNICIPIOS_MENU = [
    ("Quimbaya", "QUIMBAYA"),
    ("Montenegro", "MONTENEGRO"),
    ("La Tebaida", "LA TEBAIDA"),
    ("Circasia", "CIRCASIA"),
    ("Filandia", "FILANDIA"),
    ("Salento", "SALENTO"),
    ("Calarcá", "CALARCA"),
    ("Buenavista", "BUENAVISTA"),
    ("Génova", "GENOVA"),
    ("Córdoba", "CORDOBA"),
    ("Pijao", "PIJAO"),
]

HOJAS_AUXILIARES = {"CLASIFICADORES", "ESTADISTICO", "LISTADO COMUNAS"}

RANGOS_EDAD = ["14-28", "29-59", "60-69", "70+"]
GENEROS_ORDEN = ["Femenino", "Masculino", "LGBTI", "NS/NR"]

COLORES_ESTADO = {
    "ACTIVO": "#2A803B", # Verde Gobernación
    "INACTIVO": "#E74C3C", # Rojo
    "OTRO": "#F39C12", # Amarillo
    "SIN_DATO": "#BDC3C7",
}

COLORES_GENERO = {
    "Femenino": "#5C2D91", # Morado
    "Masculino": "#2A803B", # Verde
    "LGBTI": "#FFC72C", # Amarillo
    "NS/NR": "#95A5A6",
}

COLORES_SECTOR = {
    "Urbano": "#2980B9", # Azul
    "Rural": "#2A803B", # Verde Gobernación
    "Sin clasificar": "#FFC72C", # Amarillo
    "Sin dato": "#95A5A6",
}

ROLES_DIGNATARIOS = [
    "PRESIDENTE",
    "VICEPRESIDENTE",
    "TESORERO",
    "SECRETARIO",
    "FISCAL",
    "SUPLENTE_FISCAL",
    "CONCILIADOR_01",
    "CONCILIADOR_02",
    "CONCILIADOR_03",
    "DELEGADO_1",
    "DELEGADO_SUPLENTE_1",
    "DELEGADO_2",
    "DELEGADO_SUPLENTE_2",
    "DELEGADO_3",
    "DELEGADO_SUPLENTE_3",
    "DELEGADO_4",
    "DELEGADO_SUPLENTE_4",
    "DELEGADO_SUPLENTE",
    "REPRESENTANTE_CT_1",
    "REPRESENTANTE_CT_2",
    "REPRESENTANTE_CT_3",
    "REPRESENTANTE_CT_4",
    "REPRESENTANTE_CT_5",
    "REPRESENTANTE_CT_6",
    "REPRESENTANTE_CEM1",
    "REPRESENTANTE_CEM2",
    "REPRESENTANTE_CEM3",
]

ETIQUETAS_ROL = {
    "PRESIDENTE": "Presidente",
    "VICEPRESIDENTE": "Vicepresidente",
    "TESORERO": "Tesorero",
    "SECRETARIO": "Secretario",
    "FISCAL": "Fiscal",
    "SUPLENTE_FISCAL": "Suplente fiscal",
}

SECTORES_ORDEN = ["Urbano", "Rural", "Sin clasificar", "Sin dato"]


def normalizar_texto(valor) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def clave_municipio_desde_hoja(nombre_hoja: str) -> str | None:
    hoja = normalizar_texto(nombre_hoja).replace(".", "")
    if hoja in HOJAS_AUXILIARES:
        return None
    if hoja == "TEBAIDA":
        return "LA TEBAIDA"
    for _, clave in MUNICIPIOS_MENU:
        if clave in hoja or hoja in clave:
            return clave
    return None


def municipio_coincide_estadistico(nombre: str, clave: str) -> bool:
    n = normalizar_texto(nombre).replace(".", "")
    if not n or n in ("MUNICIPIO", "TOTAL", "TOTAL ACTIVAS"):
        return False
    if clave in n or n.startswith(clave[:5]):
        return True
    if clave == "CALARCA" and "CALARCA" in n:
        return True
    if clave == "LA TEBAIDA" and ("TEBAIDA" in n or "LA TEBAIDA" in n):
        return True
    return False


def etiqueta_municipio(clave: str) -> str:
    return dict((c, e) for e, c in MUNICIPIOS_MENU).get(clave, clave.title())


def limpiar_nombre_columna(col) -> str:
    if col is None or (isinstance(col, float) and pd.isna(col)):
        return "COL_SIN_NOMBRE"
    texto = str(col).replace("\n", " ").strip().upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", "_", texto)
    return texto


def estandarizar_estado(serie: pd.Series) -> pd.Series:
    estado = serie.astype(str).str.upper().str.strip()
    estado = estado.replace({"NAN": "", "NONE": "", "0": ""})
    estado = estado.where(estado != "", "SIN_DATO")
    estado = estado.where(~estado.str.contains("ACTIVO", na=False), "ACTIVO")
    estado = estado.where(
        ~estado.str.contains("INACTIVO|INACTIVA", na=False), "INACTIVO"
    )
    estado = estado.where(estado.isin(["ACTIVO", "INACTIVO"]), "OTRO")
    return estado


def normalizar_rango_edad(valor) -> str | None:
    v = normalizar_texto(valor)
    if not v or v in ("NR", "N/A", "ND", "SIN DATO"):
        return None
    v = v.replace(" ", "")
    if "14" in v and "28" in v:
        return "14-28"
    if "29" in v and "59" in v:
        return "29-59"
    if "60" in v and "69" in v:
        return "60-69"
    if v.startswith("60") or "60-69" in v or "60 -69" in normalizar_texto(valor):
        return "60-69"
    if "70" in v:
        return "70+"
    try:
        n = int(float(valor))
        if 14 <= n <= 28:
            return "14-28"
        if 29 <= n <= 59:
            return "29-59"
        if 60 <= n <= 69:
            return "60-69"
        if n >= 70:
            return "70+"
    except (TypeError, ValueError):
        pass
    return None


def normalizar_genero(valor) -> str | None:
    v = normalizar_texto(valor)
    if not v or v in ("NR", "N/A", "ND", "0"):
        return None
    if "FEM" in v:
        return "Femenino"
    if "MASC" in v:
        return "Masculino"
    if "LGBTI" in v:
        return "LGBTI"
    if "NS" in v or "NR" in v:
        return "NS/NR"
    return "NS/NR"


def normalizar_escolaridad(valor) -> str | None:
    v = normalizar_texto(valor)
    if not v or v in ("NR", "N/A", "ND", "0", "NAN"):
        return None
    if "POSTGRADO" in v:
        return "Postgrado"
    if "PROFESIONAL" in v:
        return "Profesional"
    if "TECNOLOG" in v:
        return "Tecnólogo"
    if "TECNIC" in v:
        return "Técnico"
    if "BACHILLER" in v:
        return "Bachiller"
    if "SECUNDARIA" in v or "BASICA SECUNDARIA" in v:
        return "Básica secundaria"
    if "PRIMARIA" in v:
        return "Primaria"
    return v.title()[:40]


def normalizar_sector(valor) -> str:
    v = normalizar_texto(valor)
    if not v or v in ("NR", "N/A", "ND", "0", "NAN"):
        return "Sin dato"
    if "RURAL" in v:
        return "Rural"
    if "URB" in v:
        return "Urbano"
    if v == "ZZZ":
        return "Sin clasificar"
    return "Sin dato"


def inferir_sector_fila(fila: pd.Series) -> str:
    for col in ("SECTOR", "URBANO", "RURAL"):
        if col in fila.index and pd.notna(fila[col]):
            sector = normalizar_sector(fila[col])
            if sector != "Sin dato":
                return sector
    return "Sin dato"


def normalizar_denominacion(valor) -> str | None:
    v = normalizar_texto(valor)
    if not v or v in ("NR", "N/A", "ND", "0", "NAN"):
        return None
    if v == "BARRIO" or v.startswith("BARRIO "):
        return "Barrio"
    if v == "VEREDA" or v.startswith("VEREDA"):
        return "Vereda"
    if "URBANIZ" in v:
        return "Urbanización"
    if "JVC" in v or "J.V.C" in v:
        return "JVC"
    if "CORREGIMIENTO" in v:
        return "Corregimiento"
    if "COMUNA" in v:
        return "Comuna"
    if "JUNTA" in v or "JAC" in v:
        return "Junta comunal"
    return v.title()[:35]


def nombre_organismo(fila: pd.Series) -> str:
    for col in ("N0MBRE_DEL_ORGANISMO_COMUNAL", "N0MBRE", "DENOMINACION"):
        if col in fila.index and pd.notna(fila[col]):
            nombre = str(fila[col]).strip()
            if nombre and nombre.lower() not in ("nan", "none"):
                return nombre
    return "Sin nombre"


def mapear_bloques_roles(columnas: list[str]) -> dict[str, dict[str, str | None]]:
    """Mapea cada columna de rol a sus columnas de género, edad y escolaridad."""
    indices_rol = [(i, c) for i, c in enumerate(columnas) if c in ROLES_DIGNATARIOS]
    bloques: dict[str, dict[str, str | None]] = {}

    for idx, (inicio, rol) in enumerate(indices_rol):
        fin = indices_rol[idx + 1][0] if idx + 1 < len(indices_rol) else len(columnas)
        segmento = columnas[inicio:fin]
        genero = next((c for c in segmento if "GENERO" in c.upper()), None)
        edad = next(
            (
                c
                for c in segmento
                if c.upper().startswith("EDAD") or c in ("EDAD", "EDAD_P")
            ),
            None,
        )
        escolaridad = next((c for c in segmento if c.startswith("ESCOLARIDAD")), None)
        bloques[rol] = {
            "nombre": rol,
            "genero": genero,
            "edad": edad,
            "escolaridad": escolaridad,
        }
    return bloques


def extraer_dignatarios_por_rol(
    df: pd.DataFrame, clave_muni: str | None = None
) -> pd.DataFrame:
    """Extrae dignatarios con rol identificado (presidente vs resto)."""
    if df.empty:
        return pd.DataFrame()

    bloques = mapear_bloques_roles(list(df.columns))
    if not bloques:
        return pd.DataFrame()

    municipio = etiqueta_municipio(clave_muni) if clave_muni else "Todos"
    filas = []

    for _, row in df.iterrows():
        for rol, cols in bloques.items():
            col_nombre = cols["nombre"]
            if col_nombre not in df.columns:
                continue

            nombre = row.get(col_nombre)
            nombre_ok = (
                pd.notna(nombre)
                and str(nombre).strip() not in ("", "nan", "0", "None")
                and str(nombre).strip().upper() not in ROLES_DIGNATARIOS
            )

            genero = (
                normalizar_genero(row.get(cols["genero"]))
                if cols["genero"] and cols["genero"] in df.columns
                else None
            )
            edad = (
                normalizar_rango_edad(row.get(cols["edad"]))
                if cols["edad"] and cols["edad"] in df.columns
                else None
            )
            escolaridad = (
                normalizar_escolaridad(row.get(cols["escolaridad"]))
                if cols["escolaridad"] and cols["escolaridad"] in df.columns
                else None
            )

            if not (nombre_ok or genero or edad or escolaridad):
                continue

            grupo = "Presidente" if rol == "PRESIDENTE" else "Otros dignatarios"
            etiqueta_rol = ETIQUETAS_ROL.get(rol, rol.replace("_", " ").title())

            filas.append(
                {
                    "Municipio": municipio,
                    "Rol": etiqueta_rol,
                    "Rol_clave": rol,
                    "Grupo": grupo,
                    "Nombre": str(nombre).strip() if nombre_ok else "",
                    "Genero": genero,
                    "Edad": edad,
                    "Escolaridad": escolaridad,
                }
            )

    return pd.DataFrame(filas)


def conteo_por_grupo(dignatarios_rol: pd.DataFrame, campo: str, orden=None) -> pd.DataFrame:
    if dignatarios_rol.empty or campo not in dignatarios_rol.columns:
        return pd.DataFrame()
    tmp = dignatarios_rol.dropna(subset=[campo])
    tmp = tmp[tmp[campo].astype(str).str.strip() != ""]
    if tmp.empty:
        return pd.DataFrame()
    conteo = (
        tmp.groupby(["Grupo", campo])
        .size()
        .reset_index(name="Cantidad")
    )
    if orden:
        conteo[campo] = pd.Categorical(conteo[campo], categories=orden, ordered=True)
        conteo = conteo.sort_values([campo, "Grupo"])
    return conteo


def leer_hoja_excel(archivo, nombre_hoja: str) -> pd.DataFrame:
    df = pd.read_excel(archivo, sheet_name=nombre_hoja, engine="openpyxl")
    df.dropna(how="all", inplace=True)
    cols = []
    vistos: dict[str, int] = {}
    for c in df.columns:
        base = limpiar_nombre_columna(c)
        if base in vistos:
            vistos[base] += 1
            cols.append(f"{base}_{vistos[base]}")
        else:
            vistos[base] = 0
            cols.append(base)
    df.columns = cols
    return df


def columnas_dignatario(df: pd.DataFrame, tipo: str) -> list[str]:
    tipo = tipo.upper()
    if tipo == "EDAD":
        return [
            c
            for c in df.columns
            if c == "EDAD"
            or c.startswith("EDAD_")
            or c.startswith("EDAD.")
            or (c.endswith("_P") and "EDAD" in c)
        ]
    if tipo == "GENERO":
        return [c for c in df.columns if "GENERO" in c.upper()]
    if tipo == "ESCOLARIDAD":
        return [c for c in df.columns if c.startswith("ESCOLARIDAD")]
    return []


def extraer_valores_columnas(df: pd.DataFrame, columnas: list[str]) -> pd.Series:
    if not columnas:
        return pd.Series(dtype=object)
    partes = []
    for col in columnas:
        if col in df.columns:
            s = df[col].dropna()
            s = s.astype(str).str.strip()
            s = s[~s.isin(["", "nan", "None", "0"])]
            partes.append(s)
    if not partes:
        return pd.Series(dtype=object)
    return pd.concat(partes, ignore_index=True)


def dataframe_dignatarios(
    df: pd.DataFrame, rol: str | None = None, clave_muni: str | None = None
) -> pd.DataFrame:
    """Apila edad, género y escolaridad de todas las columnas de dignatarios."""
    filas = []
    municipio = etiqueta_municipio(clave_muni) if clave_muni else "Todos"

    edades = extraer_valores_columnas(df, columnas_dignatario(df, "EDAD"))
    for val in edades:
        rango = normalizar_rango_edad(val)
        if rango:
            filas.append({"Municipio": municipio, "Tipo": "Edad", "Categoria": rango, "Rol": rol or "Todos"})

    generos = extraer_valores_columnas(df, columnas_dignatario(df, "GENERO"))
    for val in generos:
        gen = normalizar_genero(val)
        if gen:
            filas.append({"Municipio": municipio, "Tipo": "Género", "Categoria": gen, "Rol": rol or "Todos"})

    escolaridades = extraer_valores_columnas(df, columnas_dignatario(df, "ESCOLARIDAD"))
    for val in escolaridades:
        esc = normalizar_escolaridad(val)
        if esc:
            filas.append(
                {"Municipio": municipio, "Tipo": "Escolaridad", "Categoria": esc, "Rol": rol or "Todos"}
            )

    return pd.DataFrame(filas)


def extraer_presidente_df(df: pd.DataFrame, clave_muni: str | None) -> pd.DataFrame:
    """Datos solo del bloque Presidente (columnas junto a PRESIDENTE)."""
    if "PRESIDENTE" not in df.columns:
        return pd.DataFrame()
    cols = list(df.columns)
    try:
        idx = cols.index("PRESIDENTE")
    except ValueError:
        return pd.DataFrame()
    bloque = cols[idx : idx + 10]
    sub = df[bloque].copy()
    sub.columns = [
        "Presidente",
        "Genero",
        "Edad",
        "Escolaridad",
        "Discapacidad",
        "Cedula",
        "Direccion",
        "Telefono",
        "Correo",
        "Extra",
    ][: len(bloque)]
    sub = sub[sub["Presidente"].notna() & (sub["Presidente"].astype(str).str.strip() != "")]
    sub["Municipio"] = etiqueta_municipio(clave_muni) if clave_muni else ""
    return sub


def parsear_hoja_estadistico(archivo) -> dict:
    nombre_hoja = None
    import openpyxl

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    for s in wb.sheetnames:
        if normalizar_texto(s) == "ESTADISTICO":
            nombre_hoja = s
            break
    wb.close()
    if not nombre_hoja:
        return {}

    raw = pd.read_excel(archivo, sheet_name=nombre_hoja, header=None)

    activos = []
    for i in range(3, len(raw)):
        muni = raw.iloc[i, 1]
        cant = raw.iloc[i, 2]
        if pd.isna(muni):
            continue
        m = str(muni).strip()
        if normalizar_texto(m) in ("TOTAL ACTIVAS", "TOTAL"):
            break
        try:
            activos.append({"Municipio": m, "OAC_activas": int(float(cant))})
        except (TypeError, ValueError):
            continue

    inactivos = []
    for i in range(3, len(raw)):
        muni = raw.iloc[i, 14]
        jac = raw.iloc[i, 15]
        jvc = raw.iloc[i, 16]
        if pd.isna(muni):
            continue
        m = str(muni).strip()
        if normalizar_texto(m) in ("TOTAL", "MUNICIPIO"):
            if normalizar_texto(m) == "TOTAL":
                break
            continue
        try:
            n_jac = int(float(jac)) if pd.notna(jac) else 0
            n_jvc = int(float(jvc)) if pd.notna(jvc) else 0
            inactivos.append(
                {
                    "Municipio": m,
                    "JAC_inactivas": n_jac,
                    "JVC_inactivas": n_jvc,
                    "OAC_inactivas": n_jac + n_jvc,
                }
            )
        except (TypeError, ValueError):
            continue

    jovenes = []
    for i in range(len(raw)):
        cel = raw.iloc[i, 1]
        if isinstance(cel, str) and "JOVENES" in normalizar_texto(cel):
            for j in range(i + 2, min(i + 20, len(raw))):
                muni = raw.iloc[j, 1]
                if pd.isna(muni):
                    continue
                m = str(muni).strip()
                if normalizar_texto(m) == "TOTAL":
                    break
                try:
                    jovenes.append(
                        {
                            "Municipio": m,
                            "Mujeres_dignatarias": int(float(raw.iloc[j, 2]))
                            if pd.notna(raw.iloc[j, 2])
                            else 0,
                            "Hombres": int(float(raw.iloc[j, 3]))
                            if pd.notna(raw.iloc[j, 3])
                            else 0,
                        }
                    )
                except (TypeError, ValueError):
                    pass
            break

    return {
        "oac_activas": pd.DataFrame(activos),
        "oac_inactivas": pd.DataFrame(inactivos),
        "jovenes_mujeres": pd.DataFrame(jovenes),
    }


def parsear_hoja_comunas(archivo) -> pd.DataFrame:
    import openpyxl

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    nombre = next((s for s in wb.sheetnames if normalizar_texto(s) == "LISTADO COMUNAS"), None)
    wb.close()
    if not nombre:
        return pd.DataFrame()

    raw = pd.read_excel(archivo, sheet_name=nombre, header=None)
    registros = []
    comuna_actual = None

    for i in range(len(raw)):
        row = raw.iloc[i]
        for cell in row:
            if pd.notna(cell) and "COMUNA" in normalizar_texto(cell):
                comuna_actual = str(cell).strip()
                break

        c0 = str(row[0]).strip() if pd.notna(row[0]) else ""
        c8 = str(row[8]).strip() if pd.notna(row[8]) else ""

        if c0.isdigit() and comuna_actual:
            registros.append(
                {
                    "Comuna": comuna_actual,
                    "Tipo": "OAC",
                    "No": c0,
                    "Nombre": row[2] if pd.notna(row[2]) else "",
                    "Cedula": row[3] if pd.notna(row[3]) else "",
                    "Telefono": row[4] if pd.notna(row[4]) else "",
                }
            )
        if c8.isdigit() and comuna_actual and pd.notna(row[9]):
            registros.append(
                {
                    "Comuna": comuna_actual,
                    "Tipo": "Edil",
                    "No": c8,
                    "Nombre": row[9],
                    "Cedula": row[10] if pd.notna(row[10]) else "",
                    "Telefono": row[11] if pd.notna(row[11]) else "",
                }
            )

    return pd.DataFrame(registros)


def filtrar_estadistico_por_clave(estadistico: dict, clave: str | None) -> dict:
    if not clave or not estadistico:
        return estadistico
    out = {}
    for key, df in estadistico.items():
        if df is None or df.empty:
            out[key] = df
            continue
        mask = df["Municipio"].apply(lambda m: municipio_coincide_estadistico(m, clave))
        out[key] = df[mask].copy()
    return out


def conteos_oac_desde_juntas(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "ESTADO" not in df.columns:
        return pd.DataFrame()
    rows = []
    for clave, sub in df.groupby("MUNICIPIO_CLAVE"):
        rows.append(
            {
                "Municipio": etiqueta_municipio(clave),
                "OAC_activas": int((sub["ESTADO"] == "ACTIVO").sum()),
                "OAC_inactivas": int((sub["ESTADO"] == "INACTIVO").sum()),
                "Total_OAC": len(sub),
            }
        )
    return pd.DataFrame(rows)


def enriquecer_hoja(df: pd.DataFrame, nombre_hoja: str, clave_muni: str) -> pd.DataFrame:
    df = df.copy()
    df["HOJA_EXCEL"] = nombre_hoja
    df["MUNICIPIO_CLAVE"] = clave_muni

    if "MUNICIPIO" in df.columns:
        df["MUNICIPIO"] = df["MUNICIPIO"].astype(str).str.strip()
        df.loc[df["MUNICIPIO"].isin(["", "nan", "None", "NAN"]), "MUNICIPIO"] = None
        df["MUNICIPIO"] = df["MUNICIPIO"].ffill()
    else:
        df["MUNICIPIO"] = clave_muni.title()

    if "ACTIVO" in df.columns:
        df["ESTADO"] = estandarizar_estado(df["ACTIVO"])
    else:
        df["ESTADO"] = "SIN_DATO"

    for col in ("HOMBRES", "MUJERES", "NUMERO_TOTAL_DE_AFILIADOS"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    if "NO." in df.columns:
        numeros = pd.to_numeric(df["NO."], errors="coerce")
        df = df[numeros.notna()]

    df["SECTOR_NORM"] = df.apply(inferir_sector_fila, axis=1)
    if "DENOMINACION" in df.columns:
        df["DENOMINACION_NORM"] = df["DENOMINACION"].apply(normalizar_denominacion)
    else:
        df["DENOMINACION_NORM"] = None
    df["NOMBRE_ORGANISMO"] = df.apply(nombre_organismo, axis=1)

    return df


def elegir_hojas_por_municipio(archivo) -> tuple[dict[str, str], list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    por_muni: dict[str, list[tuple[str, int]]] = {}
    auxiliares = []

    for nombre in wb.sheetnames:
        clave = clave_municipio_desde_hoja(nombre)
        if clave:
            n_cols = wb[nombre].max_column or 0
            por_muni.setdefault(clave, []).append((nombre, n_cols))
        elif normalizar_texto(nombre) in HOJAS_AUXILIARES:
            auxiliares.append(nombre)

    wb.close()
    elegidas = {clave: max(items, key=lambda x: x[1])[0] for clave, items in por_muni.items()}
    return elegidas, auxiliares


def cargar_workbook(archivo) -> dict:
    elegidas, auxiliares = elegir_hojas_por_municipio(archivo)
    frames = []
    resumen_hojas = []

    for etiqueta, clave in MUNICIPIOS_MENU:
        nombre_hoja = elegidas.get(clave)
        if not nombre_hoja:
            resumen_hojas.append(
                {"Municipio": etiqueta, "Hoja usada": "—", "Registros": 0, "Columnas": 0}
            )
            continue
        df = leer_hoja_excel(archivo, nombre_hoja)
        df = enriquecer_hoja(df, nombre_hoja, clave)
        frames.append(df)
        resumen_hojas.append(
            {
                "Municipio": etiqueta,
                "Hoja usada": nombre_hoja,
                "Registros": len(df),
                "Columnas": len(df.columns),
            }
        )

    df_global = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    dignatarios = dataframe_dignatarios(df_global, clave_muni=None)
    dignatarios_rol = extraer_dignatarios_por_rol(df_global)

    return {
        "df": df_global,
        "dignatarios": dignatarios,
        "dignatarios_rol": dignatarios_rol,
        "resumen_hojas": pd.DataFrame(resumen_hojas),
        "hojas_municipio": elegidas,
        "hojas_auxiliares": auxiliares,
        "estadistico": parsear_hoja_estadistico(archivo),
        "comunas": parsear_hoja_comunas(archivo),
    }


def grafico_barras_ordenado(df, x, y, titulo, orden=None, color_map=None, descripcion=None):
    if df.empty:
        st.info(f"Sin datos para: {titulo}")
        return
    if orden:
        df = df.copy()
        df[x] = pd.Categorical(df[x], categories=orden, ordered=True)
        df = df.sort_values(x)
    fig = px.bar(df, x=x, y=y, title=titulo, text=y, color=x, color_discrete_map=color_map or {})
    fig.update_traces(textposition="outside")
    fig.update_layout(showlegend=False, xaxis_title="", yaxis_title="Cantidad")
    st.plotly_chart(fig, use_container_width=True)
    if descripcion:
        with st.expander("Ver detalles y análisis"):
            st.info(descripcion)


def mostrar_panel_oac(df_juntas: pd.DataFrame, estadistico: dict, clave: str | None):
    st.markdown("### Organismos de Acción Comunal (OAC)")
    est = filtrar_estadistico_por_clave(estadistico, clave) if clave else estadistico

    c1, c2, c3 = st.columns(3)
    activos_est = est.get("oac_activas", pd.DataFrame())
    inactivos_est = est.get("oac_inactivas", pd.DataFrame())

    n_activos = int(activos_est["OAC_activas"].sum()) if not activos_est.empty else 0
    n_inactivos = int(inactivos_est["OAC_inactivas"].sum()) if not inactivos_est.empty else 0

    if clave and df_juntas is not None and not df_juntas.empty:
        n_activos_j = int((df_juntas["ESTADO"] == "ACTIVO").sum())
        n_inactivos_j = int((df_juntas["ESTADO"] == "INACTIVO").sum())
        if n_activos == 0:
            n_activos = n_activos_j
        if n_inactivos == 0:
            n_inactivos = n_inactivos_j

    with c1:
        st.metric("OAC activas", f"{n_activos:,}")
    with c2:
        st.metric("OAC inactivas", f"{n_inactivos:,}")
    with c3:
        st.metric("Total OAC", f"{n_activos + n_inactivos:,}")

    g1, g2 = st.columns(2)

    with g1:
        if not activos_est.empty:
            dfa = activos_est.copy()
            dfa["Municipio"] = dfa["Municipio"].str.strip()
            grafico_barras_ordenado(
                dfa, "Municipio", "OAC_activas", "OAC activas por municipio",
                descripcion="Muestra los municipios con mayor cantidad de OAC activas. Un número alto indica fuerte participación comunitaria actual."
            )
        else:
            fallback = conteos_oac_desde_juntas(df_juntas if clave is None else df_juntas)
            if not fallback.empty:
                grafico_barras_ordenado(
                    fallback, "Municipio", "OAC_activas", "OAC activas (hojas municipales)",
                    descripcion="Muestra los municipios con mayor cantidad de OAC activas, según los datos de la hoja principal."
                )
            else:
                st.info("No hay datos de OAC activas.")

    with g2:
        if not inactivos_est.empty:
            dfi = inactivos_est.copy()
            grafico_barras_ordenado(
                dfi, "Municipio", "OAC_inactivas", "OAC inactivas por municipio",
                descripcion="Muestra dónde hay concentración de OAC inactivas. Útil para dirigir esfuerzos de reactivación y asesoría desde la Gobernación."
            )
        else:
            fallback = conteos_oac_desde_juntas(df_juntas if clave is None else df_juntas)
            if not fallback.empty:
                grafico_barras_ordenado(
                    fallback, "Municipio", "OAC_inactivas", "OAC inactivas (hojas municipales)",
                    descripcion="Concentración de OAC inactivas según la hoja principal. Estos municipios podrían requerir apoyo institucional para reactivación."
                )
            else:
                st.info("No hay datos de OAC inactivas.")

    if clave is None and not activos_est.empty:
        st.subheader("Número de OAC por municipio (activas + inactivas)")
        merge = activos_est.copy()
        if not inactivos_est.empty:
            merge = merge.merge(
                inactivos_est[["Municipio", "OAC_inactivas"]],
                on="Municipio",
                how="outer",
            ).fillna(0)
            merge["Total_OAC"] = merge["OAC_activas"] + merge["OAC_inactivas"]
        else:
            merge["Total_OAC"] = merge["OAC_activas"]
        grafico_barras_ordenado(merge, "Municipio", "Total_OAC", "Total OAC por municipio", descripcion="Volumen total histórico de OAC (activas e inactivas) en cada municipio.")


def grafico_barras_agrupado(df, x, y, color, titulo, orden_x=None, color_map=None, descripcion=None):
    if df.empty:
        st.info(f"Sin datos para: {titulo}")
        return
    plot_df = df.copy()
    if orden_x:
        plot_df[x] = pd.Categorical(plot_df[x], categories=orden_x, ordered=True)
        plot_df = plot_df.sort_values(x)
    fig = px.bar(
        plot_df,
        x=x,
        y=y,
        color=color,
        barmode="group",
        title=titulo,
        text=y,
        color_discrete_map=color_map or {},
    )
    fig.update_traces(textposition="outside")
    fig.update_layout(xaxis_title="", yaxis_title="Cantidad")
    st.plotly_chart(fig, use_container_width=True)
    if descripcion:
        with st.expander("Ver detalles y análisis"):
            st.info(descripcion)


def mostrar_panel_sector_denominacion(df: pd.DataFrame, clave: str | None):
    st.markdown("### Sector y denominación de las OAC")

    if df.empty:
        st.info("Sin datos de sector o denominación.")
        return

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Urbano", f"{int((df['SECTOR_NORM'] == 'Urbano').sum()):,}")
    with c2:
        st.metric("Rural", f"{int((df['SECTOR_NORM'] == 'Rural').sum()):,}")
    with c3:
        denom_validas = df["DENOMINACION_NORM"].notna().sum() if "DENOMINACION_NORM" in df.columns else 0
        st.metric("Con denominación", f"{int(denom_validas):,}")

    g1, g2 = st.columns(2)

    with g1:
        sector = df["SECTOR_NORM"].value_counts().reset_index()
        sector.columns = ["Sector", "Cantidad"]
        grafico_barras_ordenado(
            sector,
            "Sector",
            "Cantidad",
            "Distribución por sector",
            orden=SECTORES_ORDEN,
            color_map=COLORES_SECTOR,
            descripcion="Permite visualizar si las OAC están mayormente en zonas urbanas o rurales, guiando el enfoque de las políticas públicas."
        )

    with g2:
        if "DENOMINACION_NORM" in df.columns:
            denom = df["DENOMINACION_NORM"].dropna()
            if not denom.empty:
                d = denom.value_counts().reset_index()
                d.columns = ["Denominación", "Cantidad"]
                grafico_barras_ordenado(d, "Denominación", "Cantidad", "Tipos de denominación", descripcion="Distribución de la tipología de las organizaciones (Junta Comunal, Vereda, Barrio, etc.)")
            else:
                st.info("Sin datos de denominación.")
        else:
            st.info("Columna DENOMINACIÓN no encontrada.")

    if clave is None:
        st.subheader("Sector por municipio")
        por_muni = (
            df.groupby(["MUNICIPIO_CLAVE", "SECTOR_NORM"])
            .size()
            .reset_index(name="Cantidad")
        )
        por_muni["Municipio"] = por_muni["MUNICIPIO_CLAVE"].map(
            dict((c, e) for e, c in MUNICIPIOS_MENU)
        )
        fig = px.bar(
            por_muni,
            x="Municipio",
            y="Cantidad",
            color="SECTOR_NORM",
            barmode="stack",
            title="OAC por municipio y sector",
            color_discrete_map=COLORES_SECTOR,
        )
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)

        st.subheader("Denominación por municipio")
        if "DENOMINACION_NORM" in df.columns:
            por_denom = (
                df.dropna(subset=["DENOMINACION_NORM"])
                .groupby(["MUNICIPIO_CLAVE", "DENOMINACION_NORM"])
                .size()
                .reset_index(name="Cantidad")
            )
            por_denom["Municipio"] = por_denom["MUNICIPIO_CLAVE"].map(
                dict((c, e) for e, c in MUNICIPIOS_MENU)
            )
            top_tipos = (
                por_denom.groupby("DENOMINACION_NORM")["Cantidad"]
                .sum()
                .nlargest(6)
                .index.tolist()
            )
            por_denom = por_denom[por_denom["DENOMINACION_NORM"].isin(top_tipos)]
            fig2 = px.bar(
                por_denom,
                x="Municipio",
                y="Cantidad",
                color="DENOMINACION_NORM",
                barmode="group",
                title="Principales tipos de denominación por municipio",
            )
            fig2.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig2, use_container_width=True)

    st.subheader("Organismos comunales")
    top_n = st.slider(
        "Cantidad de organismos a listar",
        min_value=5,
        max_value=30,
        value=15,
        key=f"top_org_{clave or 'global'}",
    )
    org_cols = ["NOMBRE_ORGANISMO", "SECTOR_NORM", "DENOMINACION_NORM", "ESTADO"]
    if clave is None:
        org_cols.insert(1, "MUNICIPIO_CLAVE")
    vista_org = df[[c for c in org_cols if c in df.columns]].copy()
    if "MUNICIPIO_CLAVE" in vista_org.columns:
        vista_org["Municipio"] = vista_org["MUNICIPIO_CLAVE"].map(
            dict((c, e) for e, c in MUNICIPIOS_MENU)
        )
    vista_org = vista_org.rename(
        columns={
            "NOMBRE_ORGANISMO": "Organismo",
            "SECTOR_NORM": "Sector",
            "DENOMINACION_NORM": "Denominación",
            "ESTADO": "Estado",
        }
    )
    st.dataframe(vista_org.head(top_n), hide_index=True, use_container_width=True)


def mostrar_comparacion_presidentes(dignatarios_rol: pd.DataFrame, clave: str | None):
    st.markdown("### Presidentes vs otros dignatarios")

    if dignatarios_rol.empty:
        st.info("No hay datos de dignatarios por rol.")
        return

    n_pres = int((dignatarios_rol["Grupo"] == "Presidente").sum())
    n_otros = int((dignatarios_rol["Grupo"] == "Otros dignatarios").sum())

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Presidentes", f"{n_pres:,}")
    with c2:
        st.metric("Otros dignatarios", f"{n_otros:,}")
    with c3:
        st.metric("Total dignatarios", f"{n_pres + n_otros:,}")

    g1, g2, g3 = st.columns(3)

    with g1:
        comp_edad = conteo_por_grupo(dignatarios_rol, "Edad", orden=RANGOS_EDAD)
        if not comp_edad.empty:
            grafico_barras_agrupado(
                comp_edad,
                "Edad",
                "Cantidad",
                "Grupo",
                "Edad: Presidente vs otros",
                orden_x=RANGOS_EDAD,
                descripcion="Muestra si hay una brecha generacional entre quienes ocupan la presidencia y los demás cargos dignatarios."
            )
        else:
            st.info("Sin rangos de edad por rol.")

    with g2:
        comp_gen = conteo_por_grupo(dignatarios_rol, "Genero", orden=GENEROS_ORDEN)
        if not comp_gen.empty:
            grafico_barras_agrupado(
                comp_gen,
                "Genero",
                "Cantidad",
                "Grupo",
                "Género: Presidente vs otros",
                orden_x=GENEROS_ORDEN,
                color_map={"Presidente": "#e74c3c", "Otros dignatarios": "#3498db"},
                descripcion="Ayuda a identificar la participación femenina en cargos de liderazgo (presidencia) frente a otros roles en la junta."
            )
        else:
            st.info("Sin género por rol.")

    with g3:
        comp_esc = conteo_por_grupo(dignatarios_rol, "Escolaridad")
        if not comp_esc.empty:
            top_esc = (
                comp_esc.groupby("Escolaridad")["Cantidad"]
                .sum()
                .nlargest(8)
                .index.tolist()
            )
            comp_esc = comp_esc[comp_esc["Escolaridad"].isin(top_esc)]
            grafico_barras_agrupado(
                comp_esc,
                "Escolaridad",
                "Cantidad",
                "Grupo",
                "Escolaridad: Presidente vs otros",
                descripcion="Compara el nivel educativo de los presidentes contra el resto del equipo."
            )
        else:
            st.info("Sin escolaridad por rol.")

    with st.expander("Detalle por cargo (todos los roles)", expanded=False):
        detalle = (
            dignatarios_rol.groupby("Rol")
            .size()
            .reset_index(name="Cantidad")
            .sort_values("Cantidad", ascending=False)
        )
        st.dataframe(detalle, hide_index=True, use_container_width=True)


def mostrar_panel_dignatarios(
    df: pd.DataFrame,
    dignatarios: pd.DataFrame,
    dignatarios_rol: pd.DataFrame,
    clave: str | None,
):
    st.markdown("### Dignatarios — edad, género y escolaridad")
    if clave:
        dignatarios = dataframe_dignatarios(df, clave_muni=clave)
        dignatarios_rol = extraer_dignatarios_por_rol(df, clave_muni=clave)
    elif dignatarios.empty:
        dignatarios = dataframe_dignatarios(df)

    if dignatarios.empty and dignatarios_rol.empty:
        st.info("No se encontraron datos de dignatarios en las columnas del Excel.")
        return

    mostrar_comparacion_presidentes(dignatarios_rol, clave)

    st.markdown("---")
    st.caption("Vista consolidada de todos los dignatarios (todos los cargos)")

    pres = extraer_presidente_df(df, clave)
    if not pres.empty:
        with st.expander("Detalle presidentes (muestra)", expanded=False):
            cols_show = [
                c for c in ("Presidente", "Genero", "Edad", "Escolaridad", "Municipio") if c in pres.columns
            ]
            st.dataframe(pres[cols_show].head(30), hide_index=True, use_container_width=True)

    c1, c2, c3 = st.columns(3)

    with c1:
        edad = dignatarios[dignatarios["Tipo"] == "Edad"] if not dignatarios.empty else pd.DataFrame()
        if not edad.empty:
            conteo = edad["Categoria"].value_counts().reset_index()
            conteo.columns = ["Rango", "Cantidad"]
            grafico_barras_ordenado(
                conteo, "Rango", "Cantidad", "Edad de dignatarios", orden=RANGOS_EDAD,
                descripcion="Distribución general de edades de todos los dignatarios registrados."
            )
        else:
            st.info("Sin rangos de edad registrados.")

    with c2:
        gen = dignatarios[dignatarios["Tipo"] == "Género"] if not dignatarios.empty else pd.DataFrame()
        if not gen.empty:
            conteo = gen["Categoria"].value_counts().reset_index()
            conteo.columns = ["Género", "Cantidad"]
            grafico_barras_ordenado(
                conteo,
                "Género",
                "Cantidad",
                "Género de dignatarios",
                orden=GENEROS_ORDEN,
                color_map=COLORES_GENERO,
                descripcion="Proporción total de hombres y mujeres involucrados en la acción comunal."
            )
        else:
            st.info("Sin datos de género.")

    with c3:
        esc = dignatarios[dignatarios["Tipo"] == "Escolaridad"] if not dignatarios.empty else pd.DataFrame()
        if not esc.empty:
            conteo = esc["Categoria"].value_counts().reset_index()
            conteo.columns = ["Escolaridad", "Cantidad"]
            conteo = conteo.sort_values("Cantidad", ascending=False).head(12)
            grafico_barras_ordenado(conteo, "Escolaridad", "Cantidad", "Escolaridad de dignatarios", descripcion="Niveles de escolaridad de todos los líderes, lo cual orienta estrategias de capacitación.")
        else:
            st.info("Sin datos de escolaridad.")


def mostrar_panel_auxiliares(estadistico: dict, comunas: pd.DataFrame, clave: str | None):
    st.markdown("### Datos auxiliares (hoja ESTADÍSTICO y LISTADO COMUNAS)")
    est = filtrar_estadistico_por_clave(estadistico, clave) if clave else estadistico

    jm = est.get("jovenes_mujeres", pd.DataFrame())
    if not jm.empty:
        st.subheader("Jóvenes y mujeres (ESTADÍSTICO)")
        st.dataframe(jm, hide_index=True, use_container_width=True)

    if not comunas.empty:
        st.subheader("Listado de comunas — OAC y ediles")
        st.caption(f"{len(comunas):,} registros · {comunas['Comuna'].nunique()} comunas")
        tipo = st.multiselect(
            "Filtrar por tipo",
            ["OAC", "Edil"],
            default=["OAC", "Edil"],
            key=f"comunas_tipo_{clave or 'global'}",
        )
        vista = comunas[comunas["Tipo"].isin(tipo)] if tipo else comunas
        st.dataframe(vista, hide_index=True, use_container_width=True, height=320)


def mostrar_analisis_avanzado(dignatarios_rol: pd.DataFrame, df_juntas: pd.DataFrame, clave: str | None):
    st.markdown("### Análisis Avanzado (Visualizaciones Interactivas)")

    t1, t2 = st.columns(2)
    with t1:
        if not dignatarios_rol.empty and "Edad" in dignatarios_rol.columns and "Genero" in dignatarios_rol.columns:
            st.subheader("Distribución Demográfica (Sunburst)")
            sun_df = dignatarios_rol.dropna(subset=["Edad", "Genero"])
            if not sun_df.empty:
                fig_sun = px.sunburst(
                    sun_df,
                    path=["Genero", "Edad"],
                    title="Género y Edad de Dignatarios",
                    color="Genero",
                    color_discrete_map=COLORES_GENERO
                )
                st.plotly_chart(fig_sun, use_container_width=True)
                with st.expander("Ver detalles y análisis"):
                    st.info("El gráfico Sunburst muestra la composición demográfica de los dignatarios. El círculo interno representa el género y los anillos externos subdividen cada género por rango de edad. Haz clic en las secciones para explorar a fondo.")
            else:
                st.info("Datos insuficientes para Sunburst demográfico.")
    with t2:
        if df_juntas is not None and not df_juntas.empty:
            st.subheader("Jerarquía de Juntas (Treemap)")
            tree_cols = ["MUNICIPIO", "SECTOR_NORM", "ESTADO"] if clave is None else ["SECTOR_NORM", "DENOMINACION_NORM", "ESTADO"]
            # Filter rows where all these columns are present and not null
            tree_df = df_juntas.dropna(subset=[c for c in tree_cols if c in df_juntas.columns])
            if not tree_df.empty:
                fig_tree = px.treemap(
                    tree_df,
                    path=[c for c in tree_cols if c in tree_df.columns],
                    title="Distribución de OAC por Ubicación y Estado",
                    color="ESTADO",
                    color_discrete_map=COLORES_ESTADO
                )
                st.plotly_chart(fig_tree, use_container_width=True)
                with st.expander("Ver detalles y análisis"):
                    st.info("El Treemap (mapa de árbol) ilustra el volumen de Juntas. Los cuadros más grandes indican mayor cantidad de organizaciones. Los colores diferencian entre activas e inactivas. Haz clic para acercar.")
            else:
                st.info("Datos insuficientes para Treemap.")

def generar_excel_descarga(
    df_juntas: pd.DataFrame, 
    df_dignatarios: pd.DataFrame,
    df_dignatarios_rol: pd.DataFrame,
    estadistico: dict,
    comunas: pd.DataFrame,
    clave_muni: str | None
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Juntas de Acción Comunal
        if not df_juntas.empty:
            df_juntas.to_excel(writer, sheet_name='1. Juntas', index=False)
        else:
            pd.DataFrame({"Mensaje": ["No hay datos"]}).to_excel(writer, sheet_name='1. Juntas', index=False)
            
        # 2. Dignatarios
        if not df_dignatarios.empty:
            df_dignatarios.to_excel(writer, sheet_name='2. Dignatarios', index=False)
        else:
            pd.DataFrame({"Mensaje": ["No hay datos"]}).to_excel(writer, sheet_name='2. Dignatarios', index=False)
            
        # 3. KPIs Globales
        if not df_juntas.empty:
            total_afiliados = int(df_juntas["NUMERO_TOTAL_DE_AFILIADOS"].sum()) if "NUMERO_TOTAL_DE_AFILIADOS" in df_juntas.columns else 0
            kpi_df = pd.DataFrame({
                "Métrica": ["Total Juntas en base", "OAC Activas", "OAC Inactivas", "Total Afiliados Reportados"],
                "Valor": [
                    len(df_juntas),
                    int((df_juntas['ESTADO'] == 'ACTIVO').sum()),
                    int((df_juntas['ESTADO'] == 'INACTIVO').sum()),
                    total_afiliados
                ]
            })
            kpi_df.to_excel(writer, sheet_name='3. Resumen KPIs', index=False)
            
        # 4. Análisis OAC (Activas / Inactivas por municipio)
        est = filtrar_estadistico_por_clave(estadistico, clave_muni) if clave_muni else estadistico
        activos_est = est.get("oac_activas", pd.DataFrame())
        inactivos_est = est.get("oac_inactivas", pd.DataFrame())
        if not activos_est.empty and not inactivos_est.empty:
            merge_oac = activos_est.merge(inactivos_est[["Municipio", "OAC_inactivas"]], on="Municipio", how="outer").fillna(0)
            merge_oac["Total_OAC"] = merge_oac["OAC_activas"] + merge_oac["OAC_inactivas"]
            merge_oac.to_excel(writer, sheet_name='4. OAC por Municipio', index=False)
        else:
            fallback = conteos_oac_desde_juntas(df_juntas)
            if not fallback.empty:
                fallback.to_excel(writer, sheet_name='4. OAC por Municipio', index=False)
                
        # 5. Análisis Sector y Denominación
        if not df_juntas.empty and "SECTOR_NORM" in df_juntas.columns:
            sector_counts = df_juntas["SECTOR_NORM"].value_counts().reset_index()
            sector_counts.columns = ["Sector", "Cantidad"]
            
            if "DENOMINACION_NORM" in df_juntas.columns:
                denom_counts = df_juntas["DENOMINACION_NORM"].value_counts().reset_index()
                denom_counts.columns = ["Denominación", "Cantidad"]
                sector_counts.to_excel(writer, sheet_name='5. Sector y Denominacion', index=False, startcol=0)
                denom_counts.to_excel(writer, sheet_name='5. Sector y Denominacion', index=False, startcol=3)
            else:
                sector_counts.to_excel(writer, sheet_name='5. Sector y Denominacion', index=False)

        # 6. Demografía Dignatarios
        if not df_dignatarios_rol.empty:
            comp_edad = conteo_por_grupo(df_dignatarios_rol, "Edad", orden=RANGOS_EDAD)
            comp_gen = conteo_por_grupo(df_dignatarios_rol, "Genero", orden=GENEROS_ORDEN)
            comp_esc = conteo_por_grupo(df_dignatarios_rol, "Escolaridad")
            
            if not comp_edad.empty:
                comp_edad.to_excel(writer, sheet_name='6. Demografia Dignatarios', index=False, startcol=0)
            if not comp_gen.empty:
                comp_gen.to_excel(writer, sheet_name='6. Demografia Dignatarios', index=False, startcol=4)
            if not comp_esc.empty:
                comp_esc.to_excel(writer, sheet_name='6. Demografia Dignatarios', index=False, startcol=8)
                
        # 7. Datos Auxiliares (Jóvenes y Mujeres)
        jm = est.get("jovenes_mujeres", pd.DataFrame())
        if not jm.empty:
            jm.to_excel(writer, sheet_name='7. Jovenes y Mujeres', index=False)
            
        # 8. Comunas
        if not comunas.empty:
            comunas.to_excel(writer, sheet_name='8. Comunas y Ediles', index=False)

    return output.getvalue()

def mostrar_kpis(df: pd.DataFrame):
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Juntas / OAC en base", f"{len(df):,}")
    with c2:
        st.metric("Activas (base)", f"{int((df['ESTADO'] == 'ACTIVO').sum()):,}")
    with c3:
        st.metric("Inactivas (base)", f"{int((df['ESTADO'] == 'INACTIVO').sum()):,}")
    with c4:
        total = int(df["NUMERO_TOTAL_DE_AFILIADOS"].sum()) if "NUMERO_TOTAL_DE_AFILIADOS" in df.columns else 0
        st.metric("Afiliados reportados", f"{total:,}")


def main():
    st.set_page_config(
        page_title="Dashboard JAC Quindío",
        page_icon="🇨🇴",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # Inyección de CSS premium
    st.markdown("""
        <style>
        /* Fondo principal */
        .stApp {
            background-color: #f4f6f9;
        }
        
        /* Banner Superior (Hero) */
        .hero-banner {
            background: linear-gradient(135deg, #1b4f24 0%, #2A803B 100%);
            padding: 30px;
            border-radius: 12px;
            color: white;
            margin-top: -40px;
            margin-bottom: 25px;
            box-shadow: 0 4px 15px rgba(42, 128, 59, 0.3);
            text-align: center;
        }
        .hero-banner h1 {
            color: white !important;
            font-size: 2.5rem;
            margin-bottom: 5px;
        }
        .hero-banner p {
            font-size: 1.1rem;
            opacity: 0.9;
        }
        
        /* Tarjetas de información */
        .info-card {
            background-color: white;
            padding: 25px;
            border-radius: 12px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.04);
            border-top: 4px solid #FFC72C; /* Amarillo Quindío */
            margin-bottom: 20px;
        }
        .info-card h3 {
            color: #1b4f24 !important;
            margin-top: 0;
        }
        
        /* Ocultar elementos de Streamlit por defecto si se desea más limpieza */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        
        /* Estilos para las métricas (Cards) */
        div[data-testid="stMetric"] {
            background-color: white;
            border-radius: 10px;
            padding: 15px 20px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
            border-left: 5px solid #2A803B;
        }
        
        /* Botones y subida de archivos */
        .stFileUploader {
            background-color: white;
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.04);
            border: 1px solid #e0e0e0;
        }
        </style>
    """, unsafe_allow_html=True)

    # Banner superior personalizado
    st.markdown("""
        <div class="hero-banner">
            <h1>🇨🇴 Panel de Juntas de Acción Comunal</h1>
            <p>Departamento del Quindío • Herramienta de Inteligencia de Datos</p>
        </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "📂 **Cargar archivo Excel de las Elecciones (.xlsx)**",
        type=["xlsx", "xls"],
        help="Sube el archivo departamental consolidado, ej: 'BASE DE DATOS nuevas elecciones.xlsx'"
    )

    if uploaded is None:
        # Panel informativo con estilo de tarjeta
        st.markdown("""
        <div class="info-card">
            <h3>💡 ¿Qué es esta aplicación?</h3>
            <p>Este panel de control interactivo está diseñado para analizar el estado de las <b>Organizaciones de Acción Comunal (OAC)</b> en el departamento del Quindío, permitiendo a la Gobernación tomar decisiones informadas basadas en datos reales.</p>
        </div>
        
        <div class="info-card" style="border-top-color: #2A803B;">
            <h3>⚙️ ¿Cómo funciona?</h3>
            <p>1. <b>Sube tu base de datos:</b> Utiliza el botón de arriba para cargar el archivo Excel con los datos de las juntas. Asegúrate de que contenga las hojas de los municipios y la pestaña 'ESTADISTICO'.</p>
            <p>2. <b>Procesamiento Automático:</b> El sistema limpia, estandariza y cruza la información al instante.</p>
            <p>3. <b>Explora los resultados:</b> Utiliza el menú lateral para ver el resumen global departamental o filtrar por municipio específico.</p>
        </div>
        
        <div class="info-card" style="border-top-color: #2980B9;">
            <h3>📈 ¿Qué información te mostrará?</h3>
            <ul>
                <li><b>Métricas Clave:</b> Cantidad de juntas activas e inactivas en todo el departamento.</li>
                <li><b>Distribución Demográfica:</b> Gráficas modernas e interactivas sobre género, rangos de edad y escolaridad de los líderes (Dignatarios).</li>
                <li><b>Jerarquía:</b> Detalles sobre los presidentes frente a otros roles, tipos de denominación (Barrio, Vereda, etc.) y sectores (Urbano/Rural).</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
        return

    cache_key = f"{uploaded.name}_{uploaded.size}"
    if st.session_state.get("cache_key") != cache_key:
        with st.spinner("Leyendo hojas de municipios, ESTADÍSTICO y comunas…"):
            try:
                st.session_state.datos = cargar_workbook(uploaded)
                st.session_state.cache_key = cache_key
            except Exception as exc:
                st.error(f"No se pudo procesar el archivo: {exc}")
                return

    datos = st.session_state.datos
    df = datos["df"]

    if df.empty:
        st.error("No se encontraron datos de municipios en el archivo.")
        return

    st.sidebar.header("Navegación")
    opciones = ["🌍 Vista global"] + [e for e, _ in MUNICIPIOS_MENU]
    seleccion = st.sidebar.radio("Municipio", opciones, label_visibility="collapsed")
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("Filtros Adicionales")
    filtro_estado = st.sidebar.selectbox("Estado de OAC", ["Todos", "ACTIVO", "INACTIVO"], help="Filtra las organizaciones según su estado jurídico.")
    filtro_sector = st.sidebar.selectbox("Sector", ["Todos", "Urbano", "Rural"], help="Filtra según el sector de la organización.")

    # Aplicar filtros al DataFrame principal
    if filtro_estado != "Todos":
        df = df[df["ESTADO"] == filtro_estado]
    if filtro_sector != "Todos":
        df = df[df["SECTOR_NORM"] == filtro_sector]
        
    # Re-calcular los dignatarios si hubo filtrado
    dignatarios_df = datos["dignatarios"]
    dignatarios_rol_df = datos["dignatarios_rol"]
    if filtro_estado != "Todos" or filtro_sector != "Todos":
        dignatarios_df = dataframe_dignatarios(df)
        dignatarios_rol_df = extraer_dignatarios_por_rol(df)

    st.sidebar.markdown("---")
    st.sidebar.subheader("Hojas cargadas")
    st.sidebar.dataframe(datos["resumen_hojas"], hide_index=True, use_container_width=True)

    if seleccion == "🌍 Vista global":
        st.subheader("Análisis global — departamento del Quindío")
        mostrar_kpis(df)
        st.markdown("---")
        mostrar_panel_oac(df, datos["estadistico"], clave=None)
        st.markdown("---")
        mostrar_panel_sector_denominacion(df, clave=None)
        st.markdown("---")
        mostrar_panel_dignatarios(
            df, dignatarios_df, dignatarios_rol_df, clave=None
        )
        st.markdown("---")
        mostrar_panel_auxiliares(datos["estadistico"], datos["comunas"], clave=None)

        st.markdown("---")
        mostrar_analisis_avanzado(dignatarios_rol_df, df, clave=None)

        st.markdown("---")
        st.subheader("Resumen por municipio (base de datos)")
        resumen = []
        for etiqueta, clave in MUNICIPIOS_MENU:
            sub = df[df["MUNICIPIO_CLAVE"] == clave]
            resumen.append(
                {
                    "Municipio": etiqueta,
                    "OAC en base": len(sub),
                    "Activas": int((sub["ESTADO"] == "ACTIVO").sum()),
                    "Inactivas": int((sub["ESTADO"] == "INACTIVO").sum()),
                }
            )
        st.dataframe(pd.DataFrame(resumen), hide_index=True, use_container_width=True)
        
        st.markdown("---")
        st.subheader("📥 Exportar Datos")
        excel_data = generar_excel_descarga(
            df, 
            dignatarios_df, 
            dignatarios_rol_df, 
            datos["estadistico"], 
            datos["comunas"], 
            clave_muni=None
        )
        st.download_button(
            label="📥 Descargar Reporte Global (Excel Multi-hoja)",
            data=excel_data,
            file_name="Reporte_JAC_Global.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    else:
        clave = dict((e, c) for e, c in MUNICIPIOS_MENU)[seleccion]
        df_muni = df[df["MUNICIPIO_CLAVE"] == clave].copy()
        hoja = datos["hojas_municipio"].get(clave, "")

        st.subheader(f"Análisis — {seleccion}")
        st.caption(f"Hoja: `{hoja}` · {len(df_muni):,} registros")

        if df_muni.empty:
            st.warning("No hay registros para este municipio.")
            return

        mostrar_kpis(df_muni)
        st.markdown("---")
        mostrar_panel_oac(df_muni, datos["estadistico"], clave=clave)
        st.markdown("---")
        mostrar_panel_sector_denominacion(df_muni, clave=clave)
        st.markdown("---")
        mostrar_panel_dignatarios(
            df_muni,
            dataframe_dignatarios(df_muni, clave_muni=clave),
            extraer_dignatarios_por_rol(df_muni, clave_muni=clave),
            clave=clave,
        )
        st.markdown("---")
        mostrar_panel_auxiliares(datos["estadistico"], datos["comunas"], clave=clave)

        st.markdown("---")
        mostrar_analisis_avanzado(
            extraer_dignatarios_por_rol(df_muni, clave_muni=clave),
            df_muni,
            clave=clave
        )

        st.markdown("---")
        st.subheader("Tabla de juntas")
        columnas_vista = [
            c
            for c in (
                "NO.",
                "NOMBRE_ORGANISMO",
                "DENOMINACION",
                "DENOMINACION_NORM",
                "N0MBRE_DEL_ORGANISMO_COMUNAL",
                "MUNICIPIO",
                "ESTADO",
                "SECTOR_NORM",
                "SECTOR",
                "URBANO",
                "NUMERO_TOTAL_DE_AFILIADOS",
                "HOMBRES",
                "MUJERES",
            )
            if c in df_muni.columns
        ]
        st.dataframe(
            df_muni[columnas_vista] if columnas_vista else df_muni,
            use_container_width=True,
            height=400,
            hide_index=True,
        )
        
        st.markdown("---")
        st.subheader("📥 Exportar Datos del Municipio")
        dignatarios_muni_df = dataframe_dignatarios(df_muni, clave_muni=clave)
        dignatarios_rol_muni_df = extraer_dignatarios_por_rol(df_muni, clave_muni=clave)
        excel_data_muni = generar_excel_descarga(
            df_muni, 
            dignatarios_muni_df, 
            dignatarios_rol_muni_df, 
            datos["estadistico"], 
            datos["comunas"], 
            clave_muni=clave
        )
        st.download_button(
            label=f"📥 Descargar Reporte de {seleccion} (Excel Multi-hoja)",
            data=excel_data_muni,
            file_name=f"Reporte_JAC_{seleccion.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.success(f"Archivo **{uploaded.name}** cargado con éxito.")


if __name__ == "__main__":
    main()
